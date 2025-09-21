import pandas as pd
from openpyxl import load_workbook

COL_ANIMAL_ID = "D"

DATA_FOLDER = "./claw_segmentation/data/"
RESULT_FILE = DATA_FOLDER + "ergebnisse/ergebnisse.csv"
RESULT_FILE2 = DATA_FOLDER + "ergebnisse/ergebnisse2.csv"
REL_AREA_COL = "Relative Klauenfläche (=Aussen/Innen)"

def get_birth_dates(excel_file):
    COL_BIRTH_DATE = "F"

    wb = load_workbook(excel_file, data_only=True)
    start_row = 6
    end_row = 100
    sh = wb["Daten"]
    birth_dates = []
    for row_num in range(start_row, end_row):
        animal_id = sh[COL_ANIMAL_ID + str(row_num)].value
        birth_date = sh[COL_BIRTH_DATE + str(row_num)].value  # Format dd/mm/yyyy
        birth_date = pd.to_datetime(birth_date, format="%d/%m/%Y")
        row = {"Ohrmarkennummer": animal_id, "Geburtsdatum": birth_date}
        birth_dates.append(row)
    return pd.DataFrame(birth_dates)

def get_weight_measures(excel_file):
    COL_WEIGHT = "BR"
    COL_WEIGHT_DATE = "BK"
    COL_LINEAR_LEG_STAND = "BM"
    COL_TOEAXIS = "BN"
    COL_GROUND_WIDTH = "BO"
    COL_TRACHTEN_WIDTH_RIGHT = "BP"
    COL_TRACHTEN_WIDTH_LEFT = "BQ"

    wb = load_workbook(excel_file, data_only=True)
    start_row = 6
    end_row = 100
    sh = wb["Daten"]
    weights = []
    for row in range(start_row, end_row):
        animal_id = sh[COL_ANIMAL_ID + str(row)].value
        weight = sh[COL_WEIGHT + str(row)].value
        if weight is None:
            continue
        weight_date = sh[COL_WEIGHT_DATE + str(row)].value
        weight_date = pd.to_datetime(weight_date, format="%d/%m/%Y")
        leg_stand = sh[COL_LINEAR_LEG_STAND + str(row)].value
        toeaxis = sh[COL_TOEAXIS + str(row)].value
        ground_width = sh[COL_GROUND_WIDTH + str(row)].value
        trachten_width_right = sh[COL_TRACHTEN_WIDTH_RIGHT + str(row)].value
        trachten_width_left = sh[COL_TRACHTEN_WIDTH_LEFT + str(row)].value

        row = {
                "Ohrmarkennummer": animal_id,
                "Gewicht": weight,
                "Gewichtsdatum": weight_date,
                "Lineare Beinstellung": leg_stand,
                "Zehenachse": toeaxis,
                "Bodenweite": ground_width,
                "Trachtenhoehe rechts": trachten_width_right,
                "Trachtenhoehe links": trachten_width_left,
        }
        weights.append(row)
    return pd.DataFrame(weights)



pd.read_csv(RESULT_FILE)
res_df = pd.read_csv(RESULT_FILE, dtype=str)
res_df[REL_AREA_COL] = res_df[REL_AREA_COL].astype(float)
res_df["Datum"] = pd.to_datetime(res_df["Datum"], format="%Y_%m_%d")

excel_file = DATA_FOLDER + "Klauenerziehung_Urliste.xlsx"
birth_dates_df = get_birth_dates(excel_file).drop_duplicates(ignore_index=True)
weight_df = get_weight_measures(excel_file)

ear_tags = res_df["Ohrmarkennummer"].unique()
# Add group data to the dataframe
res_df = res_df.join(
    birth_dates_df.set_index("Ohrmarkennummer", drop=True),
    on="Ohrmarkennummer",
    how="left",
)
res_df["Alter [Tage]"] = res_df["Datum"] - res_df["Geburtsdatum"]
res_df["Alter [Tage]"] = res_df["Alter [Tage]"].dt.days.astype(int, errors="ignore")
res_df.to_csv(RESULT_FILE2, index=False)
